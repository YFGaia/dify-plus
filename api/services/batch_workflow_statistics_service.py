"""
批量工作流统计服务 - 生成专业的Excel报表
"""
from datetime import datetime, timedelta
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import text
from sqlalchemy.orm import Session

from app_factory import create_app
from extensions.ext_database import db


class BatchWorkflowStatisticsService:
    """批量工作流统计服务"""

    @staticmethod
    def get_today_app_usage_stats(session: Session | None = None) -> list[dict[str, Any]]:
        """
        获取今天各个APP的使用统计（按使用次数排序）
        
        Returns:
            list[dict]: 包含app_id, app_name, usage_count的列表，按使用次数降序
        """
        if session is None:
            session = db.session
            
        # 获取今天的开始时间（00:00:00）
        today_start = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        
        # SQL查询：统计今天各个APP的使用次数
        query = text("""
            SELECT 
                c.id as app_id,
                c.name as app_name,
                COUNT(a.id) as usage_count,
                SUM(a.total_rows) as total_rows,
                SUM(a.processed_rows) as processed_rows,
                SUM(a.error_count) as error_count
            FROM batch_workflows_extend as a
            INNER JOIN installed_apps as b ON a.installed_id::uuid = b.id
            INNER JOIN apps as c ON b.app_id = c.id
            WHERE a.created_at >= :today_start
            GROUP BY c.id, c.name
            ORDER BY usage_count DESC
        """)
        
        result = session.execute(query, {"today_start": today_start})
        
        stats = []
        for row in result:
            stats.append({
                "app_id": row.app_id,
                "app_name": row.app_name,
                "usage_count": row.usage_count,
                "total_rows": row.total_rows or 0,
                "processed_rows": row.processed_rows or 0,
                "error_count": row.error_count or 0,
            })
        
        return stats

    @staticmethod
    def get_hourly_execution_stats(session: Session | None = None, hours: int = 24) -> list[dict[str, Any]]:
        """
        获取按小时统计的执行情况
        
        Args:
            session: 数据库会话
            hours: 统计最近多少小时，默认24小时
            
        Returns:
            list[dict]: 包含时间段和执行数量的列表
        """
        if session is None:
            session = db.session
            
        # 获取起始时间
        start_time = datetime.now() - timedelta(hours=hours)
        
        # SQL查询：按小时统计执行中的任务数
        query = text("""
            SELECT 
                DATE_TRUNC('hour', a.created_at) as hour_period,
                COUNT(DISTINCT a.id) as total_count,
                COUNT(DISTINCT CASE WHEN a.status = 'processing' THEN a.id END) as processing_count,
                COUNT(DISTINCT CASE WHEN a.status = 'completed' THEN a.id END) as completed_count,
                COUNT(DISTINCT CASE WHEN a.status = 'failed' THEN a.id END) as failed_count,
                COUNT(DISTINCT CASE WHEN a.status = 'pending' THEN a.id END) as pending_count,
                SUM(a.total_rows) as total_rows,
                SUM(a.processed_rows) as processed_rows
            FROM batch_workflows_extend as a
            WHERE a.created_at >= :start_time
            GROUP BY DATE_TRUNC('hour', a.created_at)
            ORDER BY hour_period DESC
        """)
        
        result = session.execute(query, {"start_time": start_time})
        
        stats = []
        for row in result:
            stats.append({
                "hour_period": row.hour_period.strftime("%Y-%m-%d %H:00:00"),
                "total_count": row.total_count,
                "processing_count": row.processing_count or 0,
                "completed_count": row.completed_count or 0,
                "failed_count": row.failed_count or 0,
                "pending_count": row.pending_count or 0,
                "total_rows": row.total_rows or 0,
                "processed_rows": row.processed_rows or 0,
            })
        
        return stats

    @staticmethod
    def get_user_batch_stats(session: Session | None = None) -> list[dict[str, Any]]:
        """
        获取今天各用户的批量处理统计
        
        Returns:
            list[dict]: 包含用户信息和统计数据的列表
        """
        if session is None:
            session = db.session
            
        # 获取今天的开始时间
        today_start = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        
        # SQL查询：统计各用户今天的批量处理情况
        # batch_workflows_extend.user_id 对应 sys_users.id (uint类型)
        query = text("""
            SELECT 
                su.id as account_id,
                COALESCE(su.nick_name, su.username) as account_name,
                su.email as account_email,
                COUNT(a.id) as batch_count,
                SUM(a.total_rows) as total_rows,
                SUM(a.processed_rows) as processed_rows,
                SUM(a.error_count) as error_count,
                COUNT(DISTINCT a.installed_id) as app_count
            FROM batch_workflows_extend as a
            INNER JOIN sys_users as su ON a.user_id = su.id
            WHERE a.created_at >= :today_start
            GROUP BY su.id, su.nick_name, su.username, su.email
            ORDER BY batch_count DESC
        """)
        
        result = session.execute(query, {"today_start": today_start})
        
        stats = []
        for row in result:
            stats.append({
                "account_id": row.account_id,
                "account_name": row.account_name,
                "account_email": row.account_email,
                "batch_count": row.batch_count,
                "total_rows": row.total_rows or 0,
                "processed_rows": row.processed_rows or 0,
                "error_count": row.error_count or 0,
                "app_count": row.app_count or 0,
            })
        
        return stats

    @staticmethod
    def get_current_executing_stats(session: Session | None = None) -> dict[str, Any]:
        """
        获取当前正在执行的批量工作流统计
        
        Returns:
            dict: 当前执行状态的统计信息
        """
        if session is None:
            session = db.session
            
        # SQL查询：获取当前执行状态统计
        query = text("""
            SELECT 
                COUNT(DISTINCT a.id) as processing_workflows,
                COUNT(DISTINCT a.user_id) as active_users,
                COUNT(DISTINCT a.installed_id) as active_apps,
                SUM(a.total_rows - a.processed_rows) as pending_rows,
                SUM(a.processed_rows) as completed_rows
            FROM batch_workflows_extend as a
            WHERE a.status IN ('processing', 'pending')
        """)
        
        result = session.execute(query).fetchone()
        
        return {
            "processing_workflows": result.processing_workflows or 0,
            "active_users": result.active_users or 0,
            "active_apps": result.active_apps or 0,
            "pending_rows": result.pending_rows or 0,
            "completed_rows": result.completed_rows or 0,
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }

    @staticmethod
    def get_app_hourly_distribution(
        app_id: str | None = None, 
        session: Session | None = None, 
        hours: int = 24
    ) -> list[dict[str, Any]]:
        """
        获取指定APP（或所有APP）的小时级别分布统计
        
        Args:
            app_id: APP ID，如果为None则统计所有APP
            session: 数据库会话
            hours: 统计最近多少小时
            
        Returns:
            list[dict]: 小时级别的统计数据
        """
        if session is None:
            session = db.session
            
        start_time = datetime.now() - timedelta(hours=hours)
        
        if app_id:
            query = text("""
                SELECT 
                    DATE_TRUNC('hour', a.created_at) as hour_period,
                    c.id as app_id,
                    c.name as app_name,
                    COUNT(a.id) as execution_count,
                    SUM(a.total_rows) as total_rows,
                    SUM(a.processed_rows) as processed_rows
                FROM batch_workflows_extend as a
                INNER JOIN installed_apps as b ON a.installed_id::uuid = b.id
                INNER JOIN apps as c ON b.app_id = c.id
                WHERE a.created_at >= :start_time AND c.id = :app_id
                GROUP BY DATE_TRUNC('hour', a.created_at), c.id, c.name
                ORDER BY hour_period DESC
            """)
            result = session.execute(query, {"start_time": start_time, "app_id": app_id})
        else:
            query = text("""
                SELECT 
                    DATE_TRUNC('hour', a.created_at) as hour_period,
                    COUNT(a.id) as execution_count,
                    COUNT(DISTINCT b.app_id) as unique_apps,
                    SUM(a.total_rows) as total_rows,
                    SUM(a.processed_rows) as processed_rows
                FROM batch_workflows_extend as a
                INNER JOIN installed_apps as b ON a.installed_id::uuid = b.id
                WHERE a.created_at >= :start_time
                GROUP BY DATE_TRUNC('hour', a.created_at)
                ORDER BY hour_period DESC
            """)
            result = session.execute(query, {"start_time": start_time})
        
        stats = []
        for row in result:
            stat = {
                "hour_period": row.hour_period.strftime("%Y-%m-%d %H:00:00"),
                "execution_count": row.execution_count,
                "total_rows": row.total_rows or 0,
                "processed_rows": row.processed_rows or 0,
            }
            
            if app_id:
                stat["app_id"] = row.app_id
                stat["app_name"] = row.app_name
            else:
                stat["unique_apps"] = row.unique_apps or 0
            
            stats.append(stat)
        
        return stats

    @staticmethod
    def get_error_analysis_stats(session: Session | None = None) -> dict[str, Any]:
        """
        获取错误分析统计
        
        Returns:
            dict: 包含错误类型统计、APP错误分布、错误示例等
        """
        if session is None:
            session = db.session
            
        # 获取今天的开始时间
        today_start = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        
        # 1. 错误类型TOP10统计
        error_type_query = text("""
            SELECT 
                CASE 
                    WHEN bwt.error LIKE '%rate limit%' THEN 'Rate Limit (频率限制)'
                    WHEN bwt.error LIKE '%quota%' THEN 'Quota Exceeded (配额超限)'
                    WHEN bwt.error LIKE '%timeout%' THEN 'Timeout (超时)'
                    WHEN bwt.error LIKE '%connection%' THEN 'Connection Error (连接错误)'
                    WHEN bwt.error LIKE '%authentication%' THEN 'Authentication Error (认证错误)'
                    WHEN bwt.error LIKE '%permission%' THEN 'Permission Error (权限错误)'
                    WHEN bwt.error LIKE '%model%' THEN 'Model Error (模型错误)'
                    WHEN bwt.error LIKE '%重试超过%' THEN 'Retry Exceeded (重试超限)'
                    ELSE 'Other Error (其他错误)'
                END as error_type,
                COUNT(*) as error_count,
                MAX(bwt.error) as error_example
            FROM batch_workflow_tasks_extend as bwt
            INNER JOIN batch_workflows_extend as bw ON bwt.batch_workflow_id = bw.id
            WHERE bwt.status = 'failed' 
            AND bwt.created_at >= :today_start
            GROUP BY 
                CASE 
                    WHEN bwt.error LIKE '%rate limit%' THEN 'Rate Limit (频率限制)'
                    WHEN bwt.error LIKE '%quota%' THEN 'Quota Exceeded (配额超限)'
                    WHEN bwt.error LIKE '%timeout%' THEN 'Timeout (超时)'
                    WHEN bwt.error LIKE '%connection%' THEN 'Connection Error (连接错误)'
                    WHEN bwt.error LIKE '%authentication%' THEN 'Authentication Error (认证错误)'
                    WHEN bwt.error LIKE '%permission%' THEN 'Permission Error (权限错误)'
                    WHEN bwt.error LIKE '%model%' THEN 'Model Error (模型错误)'
                    WHEN bwt.error LIKE '%重试超过%' THEN 'Retry Exceeded (重试超限)'
                    ELSE 'Other Error (其他错误)'
                END
            ORDER BY error_count DESC
            LIMIT 10
        """)
        
        error_type_result = session.execute(error_type_query, {"today_start": today_start})
        error_types = []
        for row in error_type_result:
            error_types.append({
                "error_type": row.error_type,
                "error_count": row.error_count,
                "error_example": row.error_example[:200] + "..." if len(row.error_example) > 200 else row.error_example
            })
        
        # 2. 各APP的错误分布
        app_error_query = text("""
            SELECT 
                c.id as app_id,
                c.name as app_name,
                COUNT(bwt.id) as total_errors,
                COUNT(DISTINCT bwt.batch_workflow_id) as affected_workflows,
                COUNT(CASE WHEN bwt.error LIKE '%rate limit%' THEN 1 END) as rate_limit_errors,
                COUNT(CASE WHEN bwt.error LIKE '%quota%' THEN 1 END) as quota_errors,
                COUNT(CASE WHEN bwt.error LIKE '%重试超过%' THEN 1 END) as retry_errors,
                MAX(bwt.error) as error_example
            FROM batch_workflow_tasks_extend as bwt
            INNER JOIN batch_workflows_extend as bw ON bwt.batch_workflow_id = bw.id
            INNER JOIN installed_apps as b ON bw.installed_id::uuid = b.id
            INNER JOIN apps as c ON b.app_id = c.id
            WHERE bwt.status = 'failed' 
            AND bwt.created_at >= :today_start
            GROUP BY c.id, c.name
            ORDER BY total_errors DESC
        """)
        
        app_error_result = session.execute(app_error_query, {"today_start": today_start})
        app_errors = []
        for row in app_error_result:
            app_errors.append({
                "app_id": row.app_id,
                "app_name": row.app_name,
                "total_errors": row.total_errors,
                "affected_workflows": row.affected_workflows,
                "rate_limit_errors": row.rate_limit_errors or 0,
                "quota_errors": row.quota_errors or 0,
                "retry_errors": row.retry_errors or 0,
                "error_example": row.error_example[:200] + "..." if len(row.error_example) > 200 else row.error_example
            })
        
        # 3. 具体错误示例（最新的10个）
        error_examples_query = text("""
            SELECT 
                c.name as app_name,
                bwt.error,
                bwt.created_at,
                bwt.error_count,
                bwt.row_index
            FROM batch_workflow_tasks_extend as bwt
            INNER JOIN batch_workflows_extend as bw ON bwt.batch_workflow_id = bw.id
            INNER JOIN installed_apps as b ON bw.installed_id::uuid = b.id
            INNER JOIN apps as c ON b.app_id = c.id
            WHERE bwt.status = 'failed' 
            AND bwt.created_at >= :today_start
            ORDER BY bwt.created_at DESC
            LIMIT 10
        """)
        
        error_examples_result = session.execute(error_examples_query, {"today_start": today_start})
        error_examples = []
        for row in error_examples_result:
            error_examples.append({
                "app_name": row.app_name,
                "error": row.error,
                "created_at": row.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                "error_count": row.error_count,
                "row_index": row.row_index
            })
        
        return {
            "error_types": error_types,
            "app_errors": app_errors,
            "error_examples": error_examples,
            "total_errors": sum(et["error_count"] for et in error_types),
            "affected_apps": len(app_errors)
        }


class ExcelReportGenerator:
    """Excel报表生成器"""

    def __init__(self):
        self.service = BatchWorkflowStatisticsService()
        self.wb = Workbook()
        # 定义样式
        self.header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.title_font = Font(name="微软雅黑", size=16, bold=True, color="2F5496")
        self.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        self.center_alignment = Alignment(horizontal="center", vertical="center")
        self.left_alignment = Alignment(horizontal="left", vertical="center")

    def _apply_header_style(self, ws, row: int, max_col: int):
        """应用表头样式"""
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border

    def _apply_data_style(self, ws, start_row: int, end_row: int, max_col: int):
        """应用数据行样式"""
        for row in range(start_row, end_row + 1):
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = self.border
                if col == 1:
                    cell.alignment = self.left_alignment
                else:
                    cell.alignment = self.center_alignment

    def _auto_adjust_column_width(self, ws):
        """自动调整列宽"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def create_summary_sheet(self):
        """创建汇总页"""
        ws = self.wb.active
        ws.title = "概览汇总"

        # 标题
        ws.merge_cells("A1:F1")
        ws["A1"] = "批量工作流处理统计报表"
        ws["A1"].font = self.title_font
        ws["A1"].alignment = self.center_alignment

        ws.merge_cells("A2:F2")
        ws["A2"] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws["A2"].alignment = self.center_alignment

        # 当前执行状态
        current_stats = self.service.get_current_executing_stats()

        ws["A4"] = "当前执行状态"
        ws["A4"].font = Font(name="微软雅黑", size=14, bold=True, color="2F5496")

        headers = ["指标", "数值"]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=5, column=col, value=header)
        self._apply_header_style(ws, 5, 2)

        metrics = [
            ("正在执行的工作流数", current_stats["processing_workflows"]),
            ("活跃用户数", current_stats["active_users"]),
            ("活跃APP数", current_stats["active_apps"]),
            ("待处理行数", current_stats["pending_rows"]),
            ("已完成行数", current_stats["completed_rows"]),
        ]

        for row, (metric, value) in enumerate(metrics, start=6):
            ws.cell(row=row, column=1, value=metric)
            ws.cell(row=row, column=2, value=value)
            self._apply_data_style(ws, row, row, 2)

        self._auto_adjust_column_width(ws)

    def create_app_usage_sheet(self):
        """创建APP使用统计页"""
        ws = self.wb.create_sheet("APP使用统计")

        # 标题
        ws.merge_cells("A1:F1")
        ws["A1"] = "今天各APP使用统计"
        ws["A1"].font = self.title_font
        ws["A1"].alignment = self.center_alignment

        # 表头
        headers = ["APP名称", "使用次数", "总行数", "已处理行数", "错误数", "完成率(%)"]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=3, column=col, value=header)
        self._apply_header_style(ws, 3, len(headers))

        # 数据
        app_stats = self.service.get_today_app_usage_stats()
        for row, stat in enumerate(app_stats, start=4):
            completion_rate = (
                round((stat["processed_rows"] / stat["total_rows"]) * 100, 2)
                if stat["total_rows"] > 0
                else 0
            )
            ws.cell(row=row, column=1, value=stat["app_name"])
            ws.cell(row=row, column=2, value=stat["usage_count"])
            ws.cell(row=row, column=3, value=stat["total_rows"])
            ws.cell(row=row, column=4, value=stat["processed_rows"])
            ws.cell(row=row, column=5, value=stat["error_count"])
            ws.cell(row=row, column=6, value=completion_rate)

        if app_stats:
            self._apply_data_style(ws, 4, 3 + len(app_stats), len(headers))

            # 添加柱状图 - 使用次数
            chart1 = BarChart()
            chart1.title = "APP使用次数排行"
            chart1.style = 10
            chart1.x_axis.title = "APP"
            chart1.y_axis.title = "使用次数"

            data = Reference(ws, min_col=2, min_row=3, max_row=3 + len(app_stats))
            cats = Reference(ws, min_col=1, min_row=4, max_row=3 + len(app_stats))
            chart1.add_data(data, titles_from_data=True)
            chart1.set_categories(cats)
            chart1.height = 10
            chart1.width = 20
            ws.add_chart(chart1, "H3")

            # 添加饼图 - 使用次数占比
            if len(app_stats) <= 10:
                chart2 = PieChart()
                chart2.title = "APP使用次数占比"
                chart2.style = 10
                data = Reference(ws, min_col=2, min_row=4, max_row=3 + len(app_stats))
                cats = Reference(ws, min_col=1, min_row=4, max_row=3 + len(app_stats))
                chart2.add_data(data)
                chart2.set_categories(cats)
                chart2.height = 10
                chart2.width = 15
                ws.add_chart(chart2, "H20")

        self._auto_adjust_column_width(ws)

    def create_hourly_stats_sheet(self):
        """创建小时级别统计页"""
        ws = self.wb.create_sheet("小时执行统计")

        # 标题
        ws.merge_cells("A1:H1")
        ws["A1"] = "最近24小时执行统计"
        ws["A1"].font = self.title_font
        ws["A1"].alignment = self.center_alignment

        # 表头
        headers = ["时间段", "总数", "执行中", "已完成", "失败", "待处理", "总行数", "已处理行数"]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=3, column=col, value=header)
        self._apply_header_style(ws, 3, len(headers))

        # 数据
        hourly_stats = self.service.get_hourly_execution_stats(hours=24)
        for row, stat in enumerate(hourly_stats, start=4):
            ws.cell(row=row, column=1, value=stat["hour_period"])
            ws.cell(row=row, column=2, value=stat["total_count"])
            ws.cell(row=row, column=3, value=stat["processing_count"])
            ws.cell(row=row, column=4, value=stat["completed_count"])
            ws.cell(row=row, column=5, value=stat["failed_count"])
            ws.cell(row=row, column=6, value=stat["pending_count"])
            ws.cell(row=row, column=7, value=stat["total_rows"])
            ws.cell(row=row, column=8, value=stat["processed_rows"])

        if hourly_stats:
            self._apply_data_style(ws, 4, 3 + len(hourly_stats), len(headers))

            # 添加折线图 - 执行趋势
            chart = LineChart()
            chart.title = "执行数量趋势"
            chart.style = 10
            chart.x_axis.title = "时间"
            chart.y_axis.title = "数量"

            data = Reference(
                ws, min_col=2, min_row=3, max_col=6, max_row=3 + len(hourly_stats)
            )
            cats = Reference(ws, min_col=1, min_row=4, max_row=3 + len(hourly_stats))
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 12
            chart.width = 25
            ws.add_chart(chart, "J3")

        self._auto_adjust_column_width(ws)

    def create_user_stats_sheet(self):
        """创建用户统计页"""
        ws = self.wb.create_sheet("用户统计")

        # 标题
        ws.merge_cells("A1:G1")
        ws["A1"] = "今天用户批量处理统计"
        ws["A1"].font = self.title_font
        ws["A1"].alignment = self.center_alignment

        # 表头
        headers = ["用户名", "邮箱", "批次数", "总行数", "已处理行数", "错误数", "使用APP数"]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=3, column=col, value=header)
        self._apply_header_style(ws, 3, len(headers))

        # 数据
        user_stats = self.service.get_user_batch_stats()
        for row, stat in enumerate(user_stats, start=4):
            ws.cell(row=row, column=1, value=stat["account_name"])
            ws.cell(row=row, column=2, value=stat["account_email"])
            ws.cell(row=row, column=3, value=stat["batch_count"])
            ws.cell(row=row, column=4, value=stat["total_rows"])
            ws.cell(row=row, column=5, value=stat["processed_rows"])
            ws.cell(row=row, column=6, value=stat["error_count"])
            ws.cell(row=row, column=7, value=stat["app_count"])

        if user_stats:
            self._apply_data_style(ws, 4, 3 + len(user_stats), len(headers))

            # 添加柱状图 - 用户批次数排行
            chart = BarChart()
            chart.title = "用户批次数排行 TOP 10"
            chart.style = 10
            chart.x_axis.title = "用户"
            chart.y_axis.title = "批次数"

            max_rows = min(10, len(user_stats))
            data = Reference(ws, min_col=3, min_row=3, max_row=3 + max_rows)
            cats = Reference(ws, min_col=1, min_row=4, max_row=3 + max_rows)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 10
            chart.width = 20
            ws.add_chart(chart, "I3")

        self._auto_adjust_column_width(ws)

    def create_error_analysis_sheet(self):
        """创建错误分析页"""
        ws = self.wb.create_sheet("错误分析")

        # 标题
        ws.merge_cells("A1:H1")
        ws["A1"] = "今天错误分析统计"
        ws["A1"].font = self.title_font
        ws["A1"].alignment = self.center_alignment

        # 获取错误分析数据
        error_stats = self.service.get_error_analysis_stats()

        # 1. 错误类型TOP10统计
        ws["A3"] = "错误类型TOP10统计"
        ws["A3"].font = Font(name="微软雅黑", size=14, bold=True, color="2F5496")

        headers = ["错误类型", "错误次数", "错误示例"]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=4, column=col, value=header)
        self._apply_header_style(ws, 4, len(headers))

        for row, error_type in enumerate(error_stats["error_types"], start=5):
            ws.cell(row=row, column=1, value=error_type["error_type"])
            ws.cell(row=row, column=2, value=error_type["error_count"])
            ws.cell(row=row, column=3, value=error_type["error_example"])

        if error_stats["error_types"]:
            self._apply_data_style(ws, 5, 4 + len(error_stats["error_types"]), len(headers))

            # 添加饼图 - 错误类型分布
            if len(error_stats["error_types"]) <= 10:
                chart1 = PieChart()
                chart1.title = "错误类型分布"
                chart1.style = 10
                data = Reference(ws, min_col=2, min_row=4, max_row=4 + len(error_stats["error_types"]))
                cats = Reference(ws, min_col=1, min_row=5, max_row=4 + len(error_stats["error_types"]))
                chart1.add_data(data)
                chart1.set_categories(cats)
                chart1.height = 10
                chart1.width = 15
                ws.add_chart(chart1, "E4")

        # 2. 各APP错误分布
        start_row = 4 + len(error_stats["error_types"]) + 3
        ws.cell(row=start_row, column=1, value="各APP错误分布")
        ws.cell(row=start_row, column=1).font = Font(name="微软雅黑", size=14, bold=True, color="2F5496")

        app_headers = ["APP名称", "总错误数", "受影响工作流", "频率限制", "配额超限", "重试超限", "错误示例"]
        for col, header in enumerate(app_headers, start=1):
            ws.cell(row=start_row + 1, column=col, value=header)
        self._apply_header_style(ws, start_row + 1, len(app_headers))

        for row, app_error in enumerate(error_stats["app_errors"], start=start_row + 2):
            ws.cell(row=row, column=1, value=app_error["app_name"])
            ws.cell(row=row, column=2, value=app_error["total_errors"])
            ws.cell(row=row, column=3, value=app_error["affected_workflows"])
            ws.cell(row=row, column=4, value=app_error["rate_limit_errors"])
            ws.cell(row=row, column=5, value=app_error["quota_errors"])
            ws.cell(row=row, column=6, value=app_error["retry_errors"])
            ws.cell(row=row, column=7, value=app_error["error_example"])

        if error_stats["app_errors"]:
            self._apply_data_style(ws, start_row + 2, start_row + 1 + len(error_stats["app_errors"]), len(app_headers))

            # 添加柱状图 - APP错误排行
            chart2 = BarChart()
            chart2.title = "APP错误数量排行"
            chart2.style = 10
            chart2.x_axis.title = "APP"
            chart2.y_axis.title = "错误数量"

            max_rows = min(10, len(error_stats["app_errors"]))
            data = Reference(ws, min_col=2, min_row=start_row + 1, max_row=start_row + 1 + max_rows)
            cats = Reference(ws, min_col=1, min_row=start_row + 2, max_row=start_row + 1 + max_rows)
            chart2.add_data(data, titles_from_data=True)
            chart2.set_categories(cats)
            chart2.height = 10
            chart2.width = 20
            ws.add_chart(chart2, "I" + str(start_row + 1))

        # 3. 具体错误示例
        examples_start_row = start_row + 2 + len(error_stats["app_errors"]) + 3
        ws.cell(row=examples_start_row, column=1, value="最新错误示例")
        ws.cell(row=examples_start_row, column=1).font = Font(name="微软雅黑", size=14, bold=True, color="2F5496")

        example_headers = ["APP名称", "错误时间", "行索引", "重试次数", "错误详情"]
        for col, header in enumerate(example_headers, start=1):
            ws.cell(row=examples_start_row + 1, column=col, value=header)
        self._apply_header_style(ws, examples_start_row + 1, len(example_headers))

        for row, example in enumerate(error_stats["error_examples"], start=examples_start_row + 2):
            ws.cell(row=row, column=1, value=example["app_name"])
            ws.cell(row=row, column=2, value=example["created_at"])
            ws.cell(row=row, column=3, value=example["row_index"])
            ws.cell(row=row, column=4, value=example["error_count"])
            ws.cell(row=row, column=5, value=example["error"])

        if error_stats["error_examples"]:
            self._apply_data_style(ws, examples_start_row + 2, examples_start_row + 1 + len(error_stats["error_examples"]), len(example_headers))

        # 4. 错误统计汇总
        summary_start_row = examples_start_row + 2 + len(error_stats["error_examples"]) + 3
        ws.cell(row=summary_start_row, column=1, value="错误统计汇总")
        ws.cell(row=summary_start_row, column=1).font = Font(name="微软雅黑", size=14, bold=True, color="2F5496")

        summary_headers = ["指标", "数值"]
        for col, header in enumerate(summary_headers, start=1):
            ws.cell(row=summary_start_row + 1, column=col, value=header)
        self._apply_header_style(ws, summary_start_row + 1, 2)

        summary_data = [
            ("总错误数", error_stats["total_errors"]),
            ("受影响APP数", error_stats["affected_apps"]),
            ("错误类型数", len(error_stats["error_types"])),
        ]

        for row, (metric, value) in enumerate(summary_data, start=summary_start_row + 2):
            ws.cell(row=row, column=1, value=metric)
            ws.cell(row=row, column=2, value=value)
            self._apply_data_style(ws, row, row, 2)

        self._auto_adjust_column_width(ws)

    def generate_report(self, output_path: str | None = None) -> str:
        """
        生成完整的Excel报表

        Args:
            output_path: 输出文件路径，如果为None则自动生成

        Returns:
            str: 生成的文件路径
        """
        # 创建各个工作表
        self.create_summary_sheet()
        self.create_app_usage_sheet()
        self.create_hourly_stats_sheet()
        self.create_user_stats_sheet()
        self.create_error_analysis_sheet()

        # 确定输出路径
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"batch_workflow_report_{timestamp}.xlsx"

        # 保存文件
        self.wb.save(output_path)
        return output_path


def generate_batch_workflow_report(output_path: str | None = None) -> str:
    """
    生成批量工作流统计报表

    Args:
        output_path: 输出文件路径，如果为None则自动生成

    Returns:
        str: 生成的Excel文件路径

    示例:
        >>> # 生成报表到默认路径
        >>> filepath = generate_batch_workflow_report()
        >>> print(f"报表已生成: {filepath}")

        >>> # 生成报表到指定路径
        >>> filepath = generate_batch_workflow_report("/tmp/report.xlsx")
    """
    # 创建Flask应用上下文
    app = create_app()
    with app.app_context():
        generator = ExcelReportGenerator()
        filepath = generator.generate_report(output_path)
        print(f"✅ Excel报表已生成: {filepath}")
        return filepath


if __name__ == "__main__":
    # 生成Excel报表
    report_path = generate_batch_workflow_report()
    print("\n📊 批量工作流统计报表已生成")
    print(f"📁 文件路径: {report_path}")
    print("📈 报表包含以下工作表:")
    print("   1. 概览汇总 - 当前执行状态概览")
    print("   2. APP使用统计 - 各APP使用情况及图表")
    print("   3. 小时执行统计 - 24小时执行趋势")
    print("   4. 用户统计 - 用户批量处理统计")
    print("   5. 错误分析 - 错误类型TOP10、APP错误分布、具体错误示例")

